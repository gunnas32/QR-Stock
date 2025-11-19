import streamlit as st
import uuid
import qrcode
import os
import json
from datetime import datetime
from urllib.parse import urlparse, parse_qs
import smtplib
from email.message import EmailMessage

import numpy as np
import cv2
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from PIL import Image, ImageDraw, ImageFont

# ----------------------
# Config & Paths
# ----------------------
APP_BASE_URL = "https://qr-stock-twdxtxesmeefyyewte5uxy.streamlit.app/"

DATA_FILE = "inventory_data.json"
QRCODE_ROOT_DIR = "qrcodes"
REPORTS_DIR = "reports"
REPORT_FILE_NAME = "inventory_report.xlsx"

os.makedirs(QRCODE_ROOT_DIR, exist_ok=True)
os.makedirs(REPORTS_DIR, exist_ok=True)

# Email config (set these as environment variables in Streamlit Cloud)
SMTP_HOST = os.getenv("SMTP_HOST", "")
SMTP_PORT = int(os.getenv("SMTP_PORT", "587"))
SMTP_USER = os.getenv("SMTP_USER", "")
SMTP_PASSWORD = os.getenv("SMTP_PASSWORD", "")
EMAIL_FROM = os.getenv("EMAIL_FROM", SMTP_USER or "noreply@example.com")


# ----------------------
# Load / Save Inventory
# ----------------------
if os.path.exists(DATA_FILE):
    with open(DATA_FILE, "r") as f:
        inventory = json.load(f)
else:
    inventory = {}


def save_inventory():
    with open(DATA_FILE, "w") as f:
        json.dump(inventory, f, indent=4)


# ----------------------
# URL & QR Helpers
# ----------------------
def ensure_app_url() -> str:
    url = APP_BASE_URL.strip()
    if not url.endswith("/"):
        url += "/"
    return url


def generate_qr_images(item_code: str, part_name: str):
    """
    Generate QR code + labeled QR image for an item.
    qrcodes/<item_code>/<item_code>.png
    qrcodes/<item_code>/<item_code>_label.png
    Label shows:
      line 1: Part number or name
      line 2: Description: <code>
    """
    url = ensure_app_url()
    qr_data = f"{url}?code={item_code}"

    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_M,
        box_size=8,
        border=4,
    )
    qr.add_data(qr_data)
    qr.make(fit=True)
    qr_img = qr.make_image(fill_color="black", back_color="white").convert("RGB")

    item_dir = os.path.join(QRCODE_ROOT_DIR, item_code)
    os.makedirs(item_dir, exist_ok=True)

    qr_path = os.path.join(item_dir, f"{item_code}.png")
    qr_img.save(qr_path)

    # Label: QR + bigger text below
    width, height = qr_img.size
    label_height = int(height * 0.7)
    total_height = height + label_height

    label_img = Image.new("RGB", (width, total_height), "white")
    label_img.paste(qr_img, (0, 0))

    draw = ImageDraw.Draw(label_img)

    # Try larger TrueType font, fall back to default
    try:
        font = ImageFont.truetype("arial.ttf", size=26)
    except Exception:
        font = ImageFont.load_default()

    text = f"{part_name}\nDescription: {item_code}"

    try:
        bbox = draw.multiline_textbbox((0, 0), text, font=font, align="center")
        text_w = bbox[2] - bbox[0]
        text_h = bbox[3] - bbox[1]
    except AttributeError:
        lines = text.split("\n")
        text_w = max(draw.textlength(line, font=font) for line in lines)
        text_h = len(lines) * (font.size + 8)

    x = (width - text_w) // 2
    y = height + (label_height - text_h) // 2

    draw.multiline_text((x, y), text, fill="black", font=font, align="center")

    label_path = os.path.join(item_dir, f"{item_code}_label.png")
    label_img.save(label_path)

    return qr_path, label_path, qr_data


def get_qr_paths(item_code: str, part_name: str):
    """Ensure QR + label exist for an item."""
    item_dir = os.path.join(QRCODE_ROOT_DIR, item_code)
    qr_path = os.path.join(item_dir, f"{item_code}.png")
    label_path = os.path.join(item_dir, f"{item_code}_label.png")

    if not os.path.exists(qr_path) or not os.path.exists(label_path):
        qr_path, label_path, qr_data = generate_qr_images(item_code, part_name)
    else:
        qr_data = f"{ensure_app_url()}?code={item_code}"

    return qr_path, label_path, qr_data


def decode_qr_from_image(uploaded_image) -> str | None:
    """Decode QR code content from an image captured with st.camera_input."""
    if uploaded_image is None:
        return None
    file_bytes = np.frombuffer(uploaded_image.getvalue(), np.uint8)
    img = cv2.imdecode(file_bytes, cv2.IMREAD_COLOR)
    if img is None:
        return None
    detector = cv2.QRCodeDetector()
    data, points, _ = detector.detectAndDecode(img)
    if not data:
        return None
    return data


def extract_code_from_qr_data(data: str) -> str | None:
    """
    Expect data to be a URL containing ?code=XYZ, but if it's just a code, use as-is.
    """
    if not data:
        return None
    if "?" not in data:
        return data.strip()
    parsed = urlparse(data)
    qs = parse_qs(parsed.query)
    codes = qs.get("code")
    if codes:
        return codes[0]
    return None


# ----------------------
# Excel Reporting & Formatting
# ----------------------
def get_report_path() -> str:
    return os.path.join(REPORTS_DIR, REPORT_FILE_NAME)


def format_transactions_sheet(ws):
    # Headers styling
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center")
    thin = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    # Columns: Time, Part / Name, Description, Action, Qty, Person, Job, Notes
    widths = [20, 32, 32, 10, 8, 18, 18, 44]

    for col_idx, width in enumerate(widths, start=1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = width
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin

    ws.freeze_panes = "A2"

    # Style all rows
    max_row = ws.max_row
    max_col = 8
    for row in ws.iter_rows(min_row=2, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            cell.border = thin
            if cell.column in (1, 4, 5):  # Time, Action, Qty center
                cell.alignment = center

    # Color-code Action column
    for r in range(2, max_row + 1):
        cell = ws["D"][r - 1]  # Column D
        val = (cell.value or "").upper()
        if val == "IN":
            cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        elif val == "OUT":
            cell.fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
        elif val == "MANUAL":
            cell.fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")


def format_stock_sheet(ws):
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center")
    thin = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    widths = [32, 32, 10]  # Part / Name, Description, Qty

    for col_idx, width in enumerate(widths, start=1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = width
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin

    ws.freeze_panes = "A2"

    max_row = ws.max_row
    max_col = 3
    for row in ws.iter_rows(min_row=2, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            cell.border = thin
            if cell.column in (1, 2, 3):
                cell.alignment = center


def update_excel_report(
    item_code: str,
    action: str,
    qty: int,
    person: str | None,
    job: str | None,
    notes: str | None,
    ts: str,
):
    """
    Update (or create) the cumulative inventory_report.xlsx.
    Sheet 1: Transactions (all-time)
    Sheet 2: Stock (current)
    """
    report_path = get_report_path()

    if os.path.exists(report_path):
        wb = load_workbook(report_path)
    else:
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Transactions"
        ws1.append(["Time", "Part / Name", "Description", "Action", "Qty", "Person", "Job", "Notes"])
        wb.create_sheet("Stock")

    ws_tx = wb["Transactions"]
    item = inventory[item_code]

    action_label = {
        "in": "IN",
        "out": "OUT",
        "manual": "MANUAL",
    }.get(action, action.upper())

    ws_tx.append([
        ts,
        item.get("name", ""),      # Part number or name
        item_code,                 # Description label
        action_label,
        int(qty),
        person or "",
        job or "",
        notes or "",
    ])

    # Rebuild Stock sheet for up-to-date view
    if "Stock" in wb.sheetnames:
        ws_stock = wb["Stock"]
        wb.remove(ws_stock)

    ws_stock = wb.create_sheet("Stock")
    ws_stock.append(["Part / Name", "Description", "Quantity"])

    for code, it in inventory.items():
        ws_stock.append([
            it.get("name", ""),
            code,
            int(it.get("quantity", 0)),
        ])

    # Apply formatting
    format_transactions_sheet(ws_tx)
    format_stock_sheet(ws_stock)

    wb.save(report_path)


# ----------------------
# Email Alert
# ----------------------
def send_stock_alert_email(
    email_to: str,
    item_code: str,
    part_name: str,
    current_qty: int,
    threshold: int,
):
    if not (SMTP_HOST and SMTP_USER and SMTP_PASSWORD and email_to):
        return

    msg = EmailMessage()
    msg["Subject"] = f"[SVI QR-Stock] Low stock alert: {part_name} ({item_code})"
    msg["From"] = EMAIL_FROM
    msg["To"] = email_to

    body = (
        f"Dear Stock Controller,\n\n"
        f"This is an automated low stock alert from SVI QR-Stock.\n\n"
        f"Part: {part_name}\n"
        f"Description: {item_code}\n"
        f"Current quantity: {current_qty}\n"
        f"Alert threshold: {threshold}\n\n"
        f"Please review this part and replenish if necessary.\n\n"
        f"Sent: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n"
    )

    msg.set_content(body)

    try:
        with smtplib.SMTP(SMTP_HOST, SMTP_PORT) as server:
            server.starttls()
            server.login(SMTP_USER, SMTP_PASSWORD)
            server.send_message(msg)
    except Exception:
        # Fail silently to avoid breaking the app if email fails.
        pass


def maybe_send_stock_alert(item_code: str, old_qty: int, new_qty: int):
    item = inventory[item_code]
    threshold = int(item.get("alert_threshold") or 0)
    email_to = (item.get("alert_email") or "").strip()

    if threshold <= 0 or not email_to:
        return

    # Only alert when crossing from above threshold to below/equal threshold
    if old_qty > threshold >= new_qty:
        send_stock_alert_email(
            email_to,
            item_code=item_code,
            part_name=item.get("name", ""),
            current_qty=new_qty,
            threshold=threshold,
        )
        # Track last alert level
        item["last_alert_level"] = new_qty
        save_inventory()


# ----------------------
# Transactions & Item Logic
# ----------------------
def record_transaction(
    item_code: str,
    action: str,
    qty: int,
    person: str | None = None,
    job: str | None = None,
    notes: str | None = None,
):
    """Record transaction in JSON + Excel."""
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    entry = {
        "action": action,
        "qty": int(qty),
        "timestamp": ts,
    }
    if person:
        entry["person"] = person
    if job:
        entry["job"] = job
    if notes:
        entry["notes"] = notes

    inventory[item_code].setdefault("history", []).append(entry)
    save_inventory()
    update_excel_report(item_code, action, qty, person, job, notes, ts)


def show_item_view(item_code: str, header_title: str, transactions_only: bool = False):
    """
    Main item view.
    - If transactions_only=True: show only the Transaction UI (QR-scan mode).
    - Otherwise: show full tabs (Transaction, History, Edit, QR).
    """
    if item_code not in inventory:
        st.error("Part not found.")
        return

    item = inventory[item_code]
    part_name = item.get("name", "")
    qty_current = int(item.get("quantity", 0))

    st.subheader(header_title)

    # Header with stock
    header_col1, header_col2 = st.columns([3, 1])
    with header_col1:
        st.markdown(f"**Part number or name:** {part_name}")
        st.markdown(f"**Description:** `{item_code}`")
    with header_col2:
        color = "green" if qty_current > 0 else "red"
        st.markdown(
            f"<div style='text-align:center; border-radius:8px; padding:8px; "
            f"border:1px solid #DDD;'><div style='font-size:12px;color:#555;'>Stock</div>"
            f"<div style='font-size:22px;font-weight:bold;color:{color};'>{qty_current}</div></div>",
            unsafe_allow_html=True,
        )

    # ------------------- TRANSACTION BLOCK (shared) -------------------
    def render_transaction_block(show_scan_again: bool):
        st.markdown("##### Transaction")

        col_tx1, col_tx2 = st.columns(2)
        with col_tx1:
            action = st.selectbox(
                "Action",
                ["Check In", "Check Out"],
                key=f"tx_action_{item_code}",
            )
        with col_tx2:
            qty = st.number_input(
                "Quantity",
                min_value=1,
                step=1,
                key=f"tx_qty_{item_code}",
            )

        col_tx3, col_tx4 = st.columns(2)
        with col_tx3:
            person = st.text_input(
                "Person",
                key=f"tx_person_{item_code}",
            )
        with col_tx4:
            job = None
            if action == "Check Out":
                job = st.text_input(
                    "Job / Project",
                    key=f"tx_job_{item_code}",
                )

        notes = st.text_area(
            "Notes (optional)",
            key=f"tx_notes_{item_code}",
            height=60,
        )

        st.markdown("")
        tx_col1, tx_col2 = st.columns([2, 1])
        with tx_col1:
            if st.button("Apply Transaction", type="primary", key=f"tx_save_{item_code}"):
                if not person.strip():
                    st.error("Please enter the person name.")
                    return

                old_qty = qty_current
                if action == "Check In":
                    new_qty = qty_current + int(qty)
                    inventory[item_code]["quantity"] = new_qty
                    save_inventory()
                    record_transaction(
                        item_code,
                        "in",
                        qty,
                        person=person.strip(),
                        job=None,
                        notes=notes.strip() if notes else None,
                    )
                    maybe_send_stock_alert(item_code, old_qty, new_qty)
                    st.success(f"Checked IN {qty} item(s).")
                else:
                    if qty > qty_current:
                        st.error("Cannot check out more items than are in stock.")
                        return
                    new_qty = qty_current - int(qty)
                    inventory[item_code]["quantity"] = new_qty
                    save_inventory()
                    record_transaction(
                        item_code,
                        "out",
                        qty,
                        person=person.strip(),
                        job=job.strip() if job else None,
                        notes=notes.strip() if notes else None,
                    )
                    maybe_send_stock_alert(item_code, old_qty, new_qty)
                    st.success(f"Checked OUT {qty} item(s).")

                st.rerun()

        if show_scan_again:
            with tx_col2:
                if st.button("📷 Scan another item", key=f"scan_again_{item_code}"):
                    st.session_state["scan_again"] = True

            if st.session_state.get("scan_again"):
                st.markdown("#### Scan another item")
                cam_image = st.camera_input("Scan QR code of next item")
                if cam_image is not None:
                    data = decode_qr_from_image(cam_image)
                    code = extract_code_from_qr_data(data) if data else None
                    if code and code in inventory:
                        st.session_state["camera_item_code"] = code
                        st.session_state["scan_again"] = False
                        st.rerun()
                    elif code:
                        st.error(f"Scanned code '{code}' not found in inventory.")
                    else:
                        st.error("Could not read QR code from the image. Try again.")

    # ------------------- MODE: TRANSACTIONS ONLY (QR SCAN) -------------------
    if transactions_only:
        render_transaction_block(show_scan_again=True)
        return

    # ------------------- FULL TABS VIEW (DESKTOP) -------------------
    tab_tx, tab_hist, tab_edit, tab_qr = st.tabs(
        ["💼 Transaction", "📜 History", "✏️ Edit", "🖨️ QR Code"]
    )

    # Transaction tab
    with tab_tx:
        render_transaction_block(show_scan_again=False)

    # History tab
    with tab_hist:
        st.markdown("##### History")
        history = item.get("history", [])
        if not history:
            st.info("No transactions recorded yet.")
        else:
            rows = []
            for entry in reversed(history):
                rows.append({
                    "Time": entry.get("timestamp", ""),
                    "Action": {
                        "in": "IN",
                        "out": "OUT",
                        "manual": "MANUAL",
                    }.get(entry.get("action"), entry.get("action", "").upper()),
                    "Qty": entry.get("qty", 0),
                    "Person": entry.get("person", ""),
                    "Job": entry.get("job", ""),
                    "Notes": entry.get("notes", ""),
                })
            st.dataframe(rows, use_container_width=True)

    # Edit tab
    with tab_edit:
        st.markdown("##### Edit Part")

        col_e1, col_e2 = st.columns(2)
        with col_e1:
            edit_name = st.text_input(
                "Part number or name",
                value=part_name,
                key=f"edit_name_{item_code}",
            )
            edit_code = st.text_input(
                "Description",
                value=item_code,
                key=f"edit_code_{item_code}",
            )
        with col_e2:
            edit_qty = st.number_input(
                "Quantity (manual override)",
                min_value=0,
                value=qty_current,
                step=1,
                key=f"edit_qty_{item_code}",
            )

        st.markdown("##### Alert Settings")
        col_alert1, col_alert2 = st.columns(2)
        with col_alert1:
            alert_threshold = st.number_input(
                "Alert threshold (<= qty triggers email)",
                min_value=0,
                value=int(item.get("alert_threshold") or 0),
                step=1,
                key=f"alert_threshold_{item_code}",
            )
        with col_alert2:
            alert_email = st.text_input(
                "Alert email recipient",
                value=item.get("alert_email", ""),
                key=f"alert_email_{item_code}",
            )

        st.caption(
            "Note: Email alerts require SMTP settings (SMTP_HOST, SMTP_USER, SMTP_PASSWORD, EMAIL_FROM) "
            "to be configured in Streamlit Cloud."
        )

        if st.button("Save Changes", type="primary", key=f"edit_save_{item_code}"):
            code_changed = (edit_code != item_code)
            name_changed = (edit_name != part_name)
            qty_changed = (edit_qty != qty_current)

            new_code = item_code

            # Code change
            if code_changed:
                if edit_code in inventory and edit_code != item_code:
                    st.error(f"Description/code '{edit_code}' already exists. Choose another.")
                    return
                new_code = edit_code
                inventory[new_code] = inventory.pop(item_code)
                item = inventory[new_code]

            # Name change
            if name_changed:
                item["name"] = edit_name

            # Alert config changes
            item["alert_threshold"] = int(alert_threshold)
            item["alert_email"] = alert_email.strip()

            # Quantity change recorded as MANUAL
            if qty_changed:
                old_qty = qty_current
                new_qty = int(edit_qty)
                item["quantity"] = new_qty
                save_inventory()
                diff = new_qty - old_qty
                if diff != 0:
                    record_transaction(
                        new_code,
                        "manual",
                        abs(diff),
                        person="ADMIN",
                        job=None,
                        notes=f"Manual adjust from {old_qty} to {new_qty}",
                    )
                    maybe_send_stock_alert(new_code, old_qty, new_qty)
            else:
                save_inventory()

            # Regenerate QR label if name or code changed
            if code_changed or name_changed:
                _, label_path, _ = generate_qr_images(new_code, edit_name)
                st.success("Part updated and QR label regenerated.")
                st.image(label_path, caption="Updated QR label")
            else:
                st.success("Part updated.")

            st.rerun()

    # QR Code tab
    with tab_qr:
        st.markdown("##### QR Code")

        qr_path, label_path, qr_data = get_qr_paths(item_code, part_name)

        if os.path.exists(label_path):
            st.image(label_path, caption="QR Label")
        else:
            st.info("QR label will be created when needed.")

        col_q1, col_q2 = st.columns(2)
        if os.path.exists(qr_path):
            with col_q1:
                with open(qr_path, "rb") as f_qr:
                    st.download_button(
                        "Download QR (PNG)",
                        data=f_qr,
                        file_name=f"{item_code}_qr.png",
                        mime="image/png",
                        key=f"dl_qr_{item_code}",
                    )
        if os.path.exists(label_path):
            with col_q2:
                with open(label_path, "rb") as f_lbl:
                    st.download_button(
                        "Download Label (PNG)",
                        data=f_lbl,
                        file_name=f"{item_code}_label.png",
                        mime="image/png",
                        key=f"dl_label_{item_code}",
                    )

        st.caption("Scan this QR with your phone camera to open this part directly.")
        st.code(qr_data, language="text")


# ----------------------
# Streamlit App Layout
# ----------------------
st.set_page_config(page_title="SVI QR-Stock", page_icon="📦", layout="centered")

st.markdown(
    "<h2 style='margin-bottom:0.2rem;'>SVI QR-Stock</h2>"
    "<div style='color:#666;margin-bottom:1rem;'>Engineering-grade QR-based inventory for tools and parts.</div>",
    unsafe_allow_html=True,
)

# Detect QR param
scanned_code = None
qp = st.query_params
if "code" in qp and qp["code"]:
    val = qp["code"]
    scanned_code = val[0] if isinstance(val, list) else val

# Page state
# If opened via QR scan: limit functionality (no Home/Admin, item page = transaction only)
if scanned_code:
    PAGES = ["Item"]
else:
    PAGES = ["Home", "Item", "Admin"]

if "page" not in st.session_state:
    st.session_state["page"] = "Home" if not scanned_code else "Item"

# If opened via QR, default to Item page once
if scanned_code and st.session_state.get("page_init_qr") is None:
    st.session_state["page"] = "Item"
    st.session_state["page_init_qr"] = True

current_page = st.session_state["page"]
if current_page not in PAGES:
    current_page = PAGES[0]
current_index = PAGES.index(current_page)

page = st.radio(
    "Page",
    PAGES,
    index=current_index,
    horizontal=True,
    label_visibility="collapsed",
)
st.session_state["page"] = page

# ----------------------
# HOME PAGE (desktop only)
# ----------------------
if page == "Home":
    st.subheader("Overview")

    st.markdown("#### Inventory Report")
    report_path = get_report_path()
    if os.path.exists(report_path):
        with open(report_path, "rb") as f:
            st.download_button(
                "Download Inventory Report",
                data=f,
                file_name=os.path.basename(report_path),
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="dl_report_home",
            )
    else:
        st.caption("No transactions recorded yet.")

    st.markdown("#### Quick Guide")
    st.markdown(
        "- Use **Admin** to create parts and generate QR labels.\n"
        "- Attach QR labels to shelves, bins, tools, or parts.\n"
        "- Workers scan QR with phone camera to open the **Item** page in transaction-only mode.\n"
        "- Use the transaction form to check stock in or out (with person, job, notes).\n"
        "- All movements and overrides are logged into the Excel report.\n"
    )

# ----------------------
# ITEM PAGE
# ----------------------
elif page == "Item":
    st.subheader("Part")

    item_to_show = None
    transactions_only = False

    # QR-scan case: only allow transactions (phone)
    if scanned_code:
        # If user already scanned another item using camera, override
        camera_code = st.session_state.get("camera_item_code")
        effective_code = camera_code or scanned_code

        if effective_code in inventory:
            item_to_show = effective_code
            transactions_only = True
            st.caption(f"Loaded via QR scan. Mode: transactions only.")
        else:
            st.error("Scanned code not found in inventory.")

    # Manual selection (desktop)
    if not scanned_code:
        if not inventory:
            st.info("No parts in inventory yet. Use Admin to add parts.")
        else:
            st.markdown("#### Search & Select Part")

            items = []
            for code, it in sorted(inventory.items(), key=lambda x: x[1].get("name", "").lower()):
                label = f"{it.get('name','')} ({code})"
                items.append((label, code))

            search = st.text_input("Search parts (by part number or description)")
            display_items = items
            if search:
                s = search.lower()
                display_items = [
                    (lbl, code) for (lbl, code) in items if s in lbl.lower()
                ]

            labels = [lbl for (lbl, _) in display_items]
            labels_by_code = {lbl: code for (lbl, code) in display_items}

            placeholder = "[Select part]"
            options = [placeholder] + labels

            selected_label = st.selectbox("", options, index=0)

            if selected_label != placeholder:
                item_to_show = labels_by_code[selected_label]
                transactions_only = False  # full view on desktop/manual

    if item_to_show:
        show_item_view(
            item_to_show,
            header_title="Part Overview",
            transactions_only=transactions_only,
        )

# ----------------------
# ADMIN PAGE (desktop only)
# ----------------------
elif page == "Admin":
    st.subheader("Admin")

    st.markdown("#### Create New Part & QR Label")

    col_a1, col_a2 = st.columns(2)
    with col_a1:
        part_name = st.text_input("Part number or name")
    with col_a2:
        desc_input = st.text_input("Description (optional)")

    if st.button("Create Part & Generate QR Label", type="primary"):
        if not part_name:
            st.error("Please enter the part number or name.")
        else:
            item_code = desc_input.strip() or str(uuid.uuid4())[:8]

            if item_code in inventory:
                st.warning(f"Description '{item_code}' already exists. Using existing part.")
            else:
                inventory[item_code] = {
                    "name": part_name,
                    "quantity": 0,
                    "history": [],
                    "alert_threshold": 0,
                    "alert_email": "",
                }
                save_inventory()

            qr_path, label_path, qr_data = generate_qr_images(item_code, part_name)

            st.success(f"Part '{part_name}' saved with description `{item_code}`.")
            st.image(label_path, caption="QR label")

            col_d1, col_d2 = st.columns(2)
            if os.path.exists(qr_path):
                with col_d1:
                    with open(qr_path, "rb") as f_qr:
                        st.download_button(
                            "Download QR (PNG)",
                            data=f_qr,
                            file_name=f"{item_code}_qr.png",
                            mime="image/png",
                            key=f"dl_qr_new_{item_code}",
                        )
            if os.path.exists(label_path):
                with col_d2:
                    with open(label_path, "rb") as f_lbl:
                        st.download_button(
                            "Download Label (PNG)",
                            data=f_lbl,
                            file_name=f"{item_code}_label.png",
                            mime="image/png",
                            key=f"dl_label_new_{item_code}",
                        )

            st.code(qr_data, language="text")

    st.markdown("---")
    st.markdown("#### Inventory Report")

    report_path = get_report_path()
    if os.path.exists(report_path):
        with open(report_path, "rb") as f:
            st.download_button(
                "Download Inventory Report",
                data=f,
                file_name=os.path.basename(report_path),
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="dl_report_admin",
            )
    else:
        st.caption("No transactions recorded yet.")
